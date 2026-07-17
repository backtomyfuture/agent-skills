#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import sys
import glob
import json
import subprocess
import urllib.request
import urllib.error
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill

def get_env_config(env_path):
    config = {}
    if not os.path.exists(env_path):
        print(f"Error: .env file not found at {env_path}")
        sys.exit(1)
    with open(env_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if '=' in line:
                key, val = line.split('=', 1)
                config[key.strip()] = val.strip()
    return config

def get_notion_token():
    # Try environment variable first
    token = os.environ.get("NOTION_TOKEN") or os.environ.get("NOTION_API_KEY")
    if token:
        return token
    # Try macOS keychain
    try:
        res = subprocess.run(
            ['security', 'find-generic-password', '-s', 'NOTION_TOKEN', '-w'],
            capture_output=True, text=True, check=True
        )
        return res.stdout.strip()
    except Exception as e:
        print("Warning: Failed to fetch Notion token from Keychain:", e)
    return None

def query_notion_duty_dates(token, database_id, project_page_id, year, month):
    if not token or not database_id:
        print("Notion token or database ID missing, skipping Notion duty query.")
        return []

    # Filter for May 2026
    start_date = f"{year:04d}-{month:02d}-01"
    end_date = f"{year:04d}-{month:02d}-31"

    filter_data = {
        'filter': {
            'and': [
                {
                    'property': 'Project',
                    'relation': {
                        'contains': project_page_id
                    }
                },
                {
                    'property': 'Due',
                    'date': {
                        'on_or_after': start_date
                    }
                },
                {
                    'property': 'Due',
                    'date': {
                        'on_or_before': end_date
                    }
                }
            ]
        }
    }

    req = urllib.request.Request(
        f'https://api.notion.com/v1/databases/{database_id}/query',
        data=json.dumps(filter_data).encode('utf-8'),
        headers={
            'Authorization': f'Bearer {token}',
            'Notion-Version': '2022-06-28',
            'Content-Type': 'application/json'
        },
        method='POST'
    )

    duty_dates = []
    try:
        with urllib.request.urlopen(req) as res:
            data = json.loads(res.read().decode('utf-8'))
            for item in data.get('results', []):
                due_info = item['properties']['Due']['date']
                if due_info:
                    start_str = due_info['start']
                    # Extract date
                    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', start_str)
                    if m:
                        d = int(m.group(3))
                        duty_dates.append(d)
                        name = ''.join(t['plain_text'] for t in item['properties']['Name']['title'])
                        print(f"Found Notion duty task: '{name}' on day {d}")
    except Exception as e:
        print(f"Error querying Notion: {e}")
        if isinstance(e, urllib.error.HTTPError):
            print("Response body:", e.read().decode('utf-8'))

    return sorted(list(set(duty_dates)))

def query_hnaoa_documents(year, month, employee_name):
    print("Querying HNA OA documents...")
    # Year query for OA
    cmd = ['opencli', 'hnaoa', 'query', '--year', str(year), '--window', 'foreground', '-f', 'json']
    try:
        res = subprocess.run(cmd, capture_output=True, text=True, check=True)
        # Parse JSON
        docs = json.loads(res.stdout)
        print(f"Found {len(docs)} documents in total for year {year}")

        # Filter for HRM leave applications by user
        leave_docs = []
        for doc in docs:
            # We filter by reporter containing the employee name prefix (e.g. "傅强" in "傅强3")
            if doc.get('type') == 'HRM休假申请' and employee_name in doc.get('reporter', ''):
                leave_docs.append(doc)
        return leave_docs
    except Exception as e:
        print(f"Error querying HNA OA: {e}")
        return []

def get_document_details(doc, year):
    part_id = doc['part_id']
    cmd_view = ['opencli', 'hnaoa', 'view', part_id, '--year', str(year), '--window', 'foreground', '-f', 'json']
    try:
        res_view = subprocess.run(cmd_view, capture_output=True, text=True, check=True)
        details = json.loads(res_view.stdout)
        if details and len(details) > 0:
            return details[0]
    except Exception as e:
        print(f"Error getting details for doc {part_id}: {e}")
    return None

def save_document_screenshot(doc, year, output_dir, doc_detail):
    part_id = doc['part_id']
    doc_no = doc_detail.get('doc_no', 'unknown')
    title = doc_detail.get('title', '休假申请').replace('/', '_').replace('\\', '_')
    filename = f"公文_{doc_no}_{title}.png"
    filepath = os.path.join(output_dir, filename)

    print(f"Saving screenshot for document {doc_no} to {filepath}...")
    cmd_save = ['opencli', 'hnaoa', 'save', part_id, '--output', filepath, '--year', str(year), '--window', 'foreground']
    try:
        subprocess.run(cmd_save, check=True)
    except Exception as se:
        print(f"Warning: Failed to save screenshot for doc {doc_no}: {se}")

def parse_leave_from_detail(doc_detail, month):
    content = doc_detail.get('request_content', '')
    # Check if there are mentions of the month (e.g. "5月")
    month_pattern = f"{month}月"
    if month_pattern not in content:
        return []

    # Parse leave date and hours
    leaves = []
    # E.g. "5月11日（星期一）休假1天" -> day 11, 1 day
    # E.g. "5月29日（星期五）下午3点半到5点半休假2小时" -> day 29, 0.5 day (2 hours)

    # Let's search for matches
    print(f"Parsing document content: {repr(content[:150])}...")

    # Use regular expressions or custom logic
    # Find all day matches: "5月\d+日"
    matches = re.finditer(r'5月(\d+)日', content)
    for m in matches:
        day = int(m.group(1))
        # Look around the match to check for hours/days
        start_idx = max(0, m.start() - 10)
        end_idx = min(len(content), m.end() + 40)
        context = content[start_idx:end_idx].replace('\n', ' ')

        # Check type (default to 调休 since these are "补休" documents)
        leave_type = "调休"
        if "年假" in context or "年休" in context:
            leave_type = "年休假"

        # Determine duration
        days = 1.0
        if "下午" in context or "上午" in context or "2小时" in context or "4小时" in context or "半天" in context or "0.5天" in context:
            days = 0.5

        leaves.append({
            'day': day,
            'type': leave_type,
            'days': days,
            'context': context.strip()
        })
        print(f"  Parsed leave: Day {day}, Type: {leave_type}, Days: {days} (Context: '{context}')")

    return leaves

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    env_path = os.path.join(script_dir, '.env')
    config = get_env_config(env_path)

    employee_name = config.get('EMPLOYEE_NAME', '傅强')
    employee_id = config.get('EMPLOYEE_ID')
    department = config.get('DEPARTMENT')
    oa_account = config.get('OA_ACCOUNT')
    notion_page_url = config.get('NOTION_DUTY_PAGE_URL')
    notion_db_id = config.get('NOTION_DUTY_DATASOURCE_ID')

    year = 2026
    month = 5

    # Work directory
    work_dir = "/Users/jarod/Downloads/5月考勤"
    if not os.path.exists(work_dir):
        print(f"Error: Work directory {work_dir} not found.")
        sys.exit(1)

    # Scan for Excel files
    excel_files = glob.glob(os.path.join(work_dir, "*.xlsx"))
    if not excel_files:
        print(f"Error: No Excel template (.xlsx) found in {work_dir}")
        sys.exit(1)
    elif len(excel_files) > 1:
        print(f"Warning: Multiple Excel templates found: {excel_files}")

    excel_file = excel_files[0]
    print(f"Selected Excel Template: {excel_file}")

    # Step 1: Query Notion for IT duty dates
    token = get_notion_token()
    project_page_id = "2da56564-a7fe-80a8-b186-c9d4a02b06f9" # Extracted from NOTION_DUTY_PAGE_URL
    duty_days = query_notion_duty_dates(token, notion_db_id, project_page_id, year, month)
    print("IT Duty Days:", duty_days)

    # Step 2: Query HNA OA for leave documents
    leave_docs = query_hnaoa_documents(year, month, employee_name)

    # Step 3: View details, save screenshot, and parse leave
    may_leaves = []
    for doc in leave_docs:
        detail = get_document_details(doc, year)
        if detail:
            leaves = parse_leave_from_detail(detail, month)
            if leaves:
                may_leaves.extend(leaves)
                save_document_screenshot(doc, year, work_dir, detail)

    # Remove duplicate days, merging/picking max
    leave_days_map = {}
    for l in may_leaves:
        d = l['day']
        if d not in leave_days_map or l['days'] > leave_days_map[d]['days']:
            leave_days_map[d] = l

    print("Parsed May leaves summary:")
    for d, l in leave_days_map.items():
        print(f"  Day {d}: {l['type']} ({l['days']} day)")

    # Step 4: Write to Excel
    wb = openpyxl.load_workbook(excel_file, data_only=False)
    sheet = wb['异常情况明细']

    # 4.1: Identify headers and columns
    name_cols = []
    account_cols = []
    dept_cols = []
    title_cols = []
    rule_cols = []
    id_cols = []

    for col in range(1, sheet.max_column + 1):
        for r in range(1, 4):
            val = sheet.cell(r, col).value
            if val:
                val_str = str(val).strip().replace('\n', '')
                if '姓名' in val_str:
                    name_cols.append(col)
                elif '账号' in val_str:
                    account_cols.append(col)
                elif '部门' in val_str:
                    dept_cols.append(col)
                elif '职务' in val_str:
                    title_cols.append(col)
                elif '规则' in val_str:
                    rule_cols.append(col)
                elif '人事编号' in val_str:
                    id_cols.append(col)

    # Find day column mappings
    date_col_map = {}
    for col in range(39, sheet.max_column + 1):
        v2 = sheet.cell(2, col).value
        v3 = sheet.cell(3, col).value
        header = v2 or v3
        if header:
            header_str = str(header).strip()
            m = re.match(r'^(\d+)', header_str)
            if m:
                date_col_map[int(m.group(1))] = col

    # 4.2: Locate target row
    target_row = None
    # Look for name matches containing employee_name
    for r in range(4, sheet.max_row + 1):
        matched_name = False
        for c in name_cols:
            cell_val = sheet.cell(r, c).value
            if cell_val and employee_name in str(cell_val):
                matched_name = True
                break
        if matched_name:
            # Verify account or rule
            matched_account = False
            for c in account_cols:
                cell_val = sheet.cell(r, c).value
                if cell_val and oa_account in str(cell_val):
                    matched_account = True
                    break
            if matched_account:
                target_row = r
                break

    if not target_row:
        # Fallback to name match
        for r in range(4, sheet.max_row + 1):
            for c in name_cols:
                cell_val = sheet.cell(r, c).value
                if cell_val and employee_name in str(cell_val):
                    target_row = r
                    break
            if target_row:
                break

    if not target_row:
        # Create new row
        target_row = sheet.max_row + 1
        sheet.cell(target_row, 1).value = sheet.max_row - 3 # Sequence number
        print(f"Employee {employee_name} not found in Excel, creating a new row {target_row}")
    else:
        print(f"Located employee {employee_name} in row {target_row}")

    # Write values
    # B: Department
    dept_manual_col = [c for c in dept_cols if c < 10][0] if [c for c in dept_cols if c < 10] else 2
    sheet.cell(target_row, dept_manual_col).value = department

    # C: Employee ID
    id_col = id_cols[0] if id_cols else 3
    sheet.cell(target_row, id_col).value = int(employee_id) if employee_id else None

    # D: Name
    name_manual_col = [c for c in name_cols if c < 10][0] if [c for c in name_cols if c < 10] else 4
    sheet.cell(target_row, name_manual_col).value = employee_name

    # E: Personnel Type
    sheet.cell(target_row, 5).value = "行政班"

    # F: Cancel Welfare
    sheet.cell(target_row, 6).value = "否"

    # G: Required Attendance Days (19 workdays in May 2026)
    sheet.cell(target_row, 7).value = 19

    # H: Vacation Days (Sum of non-调休 leaves)
    # Since all leaves are 调休, H is 0
    non_tiaoxiu_days = sum(l['days'] for l in leave_days_map.values() if l['type'] != '调休')
    sheet.cell(target_row, 8).value = non_tiaoxiu_days

    # W: 调休⑧
    tiaoxiu_days = sum(l['days'] for l in leave_days_map.values() if l['type'] == '调休')
    sheet.cell(target_row, 23).value = tiaoxiu_days

    # X: IT duty days
    sheet.cell(target_row, 24).value = len(duty_days)

    # Fill 0 in remaining numeric columns (S:19, T:20, U:21, V:22, Y:25, Z:26, O:15)
    zero_cols = [15, 19, 20, 21, 22, 25, 26]
    for col in zero_cols:
        cell = sheet.cell(target_row, col)
        if cell.value is None or (isinstance(cell.value, str) and not cell.value.startswith('=')):
            cell.value = 0

    # Write Remark (AA)
    remark_parts = []
    for d in duty_days:
        remark_parts.append(f"5/{d} IT值班")
    for d, l in sorted(leave_days_map.items()):
        suffix = "下午 " if l['days'] == 0.5 else " "
        remark_parts.append(f"5/{d}{suffix}{l['type']}")
    remark_str = ", ".join(remark_parts)
    sheet.cell(target_row, 27).value = remark_str
    print(f"Written remark: '{remark_str}'")

    # Write date cells
    # Reset/clear previous color/values in date cells first? We only update matching dates
    for d in duty_days:
        if d in date_col_map:
            cell = sheet.cell(target_row, date_col_map[d])
            cell.value = "IT值班"
            old_font = cell.font
            cell.font = Font(name=old_font.name if old_font else None, size=old_font.size if old_font else None, bold=True, color='FF0000')

    for d, l in leave_days_map.items():
        if d in date_col_map:
            cell = sheet.cell(target_row, date_col_map[d])
            val_str = l['type']
            if l['days'] == 0.5:
                val_str += "0.5天"
            cell.value = val_str
            old_font = cell.font
            cell.font = Font(name=old_font.name if old_font else None, size=old_font.size if old_font else None, bold=True, color='FF0000')

    # Highlight the row yellow
    max_date_col = max(date_col_map.values())
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for col in range(1, max_date_col + 1):
        sheet.cell(target_row, col).fill = yellow_fill

    # Save Excel file
    wb.save(excel_file)
    print(f"Successfully saved updated考勤表 to {excel_file}")

if __name__ == '__main__':
    main()
